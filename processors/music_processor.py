"""
Music Cue Processor for DocShipper
Handles Premiere XML parsing, audio metadata extraction, and cue sheet generation
Adapted from cuemaker.py for Streamlit integration
"""

import os
import defusedxml.ElementTree as ET
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Callable
from urllib.parse import unquote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pymediainfo import MediaInfo

from utils.timecode import TimecodeHandler

logger = logging.getLogger(__name__)

# Audio file extensions recognized for cue sheet extraction
AUDIO_EXTENSIONS = {'.wav', '.mp3', '.aif', '.aiff', '.m4a', '.flac', '.aac', '.ogg'}

# Metadata fields used for richness scoring
_RICHNESS_FIELDS = ('COMPOSER', 'PERFORMER', 'ALBUM', 'GENRE', 'YEAR')


class MusicCueProcessor:
    """
    Music cue sheet processor for extracting audio metadata from Premiere XML
    and generating professional cue sheets in Excel format.
    """

    def __init__(self):
        self.tc_handler = TimecodeHandler(24)  # Default to 24fps

    @staticmethod
    def _clean_pathurl(pathurl: str) -> str:
        """Clean a Premiere XML pathurl into a usable file path."""
        clean = pathurl.replace('file://localhost', '')
        return unquote(clean)

    @staticmethod
    def _parse_timecode_info(timecode_info: str) -> tuple:
        """Parse SOURCE TIMECODES string into (tc_in, tc_out) tuple.

        Returns:
            (tc_in, tc_out) parsed from the timecode info string,
            or ("00:00:00", "00:00:00") if parsing fails.
        """
        tc_in = "00:00:00"
        tc_out = "00:00:00"
        try:
            if 'In:' in timecode_info:
                tc_in = timecode_info.split('In:')[1].split('Out:')[0].strip()
            if 'Out:' in timecode_info:
                tc_out = timecode_info.split('Out:')[1].split('Duration:')[0].strip()
        except (IndexError, ValueError):
            pass
        return tc_in, tc_out

    def _parse_xml(self, xml_path: str):
        """Parse XML file and return the ElementTree root.

        Returns:
            Root element of the parsed XML tree.

        Raises:
            ET.ParseError: If the XML file cannot be parsed.
        """
        tree = ET.parse(xml_path)
        return tree.getroot()

    def _extract_framerate_from_root(self, root) -> int:
        """Extract sequence framerate from a parsed XML root, default to 24."""
        seq = root.find('.//sequence')
        if seq is not None:
            rate = seq.find('rate')
            if rate is not None:
                timebase = rate.find('timebase')
                if timebase is not None and timebase.text.isdigit():
                    return int(timebase.text)
        return 24

    def extract_audio_paths_from_xml(self, root, filter_keywords: List[str] = None) -> Set[str]:
        """Extract audio file paths from parsed XML root using extension-based detection.

        Scans all <file> elements for <pathurl> and includes files whose extension
        matches AUDIO_EXTENSIONS. The filter_keywords parameter is accepted for
        backward compatibility but is ignored.
        """
        audio_paths = set()

        for file_elem in root.findall(".//file"):
            pathurl = file_elem.find("pathurl")
            if pathurl is not None and pathurl.text:
                clean_path = self._clean_pathurl(pathurl.text)
                ext = os.path.splitext(clean_path)[1].lower()
                if ext in AUDIO_EXTENSIONS:
                    audio_paths.add(clean_path)

        return audio_paths

    @staticmethod
    def _metadata_richness_score(row: Dict) -> int:
        """Count non-empty metadata fields for sorting. Returns 0-5."""
        return sum(1 for f in _RICHNESS_FIELDS if row.get(f, '').strip())

    def extract_clip_timecodes_from_xml(self, root, framerate: int) -> Dict[str, List[Dict]]:
        """Extract source in/out timecodes for each clip from parsed XML root."""
        clip_timecodes = {}
        self.tc_handler = TimecodeHandler(framerate)

        for clipitem in root.findall(".//clipitem"):
            file_elem = clipitem.find("file")
            if file_elem is not None:
                pathurl_elem = file_elem.find("pathurl")
                if pathurl_elem is not None and pathurl_elem.text:
                    file_path = self._clean_pathurl(pathurl_elem.text)
                    filename = os.path.basename(file_path)

                    in_elem = clipitem.find("in")
                    out_elem = clipitem.find("out")

                    if in_elem is not None and out_elem is not None:
                        try:
                            in_frames = int(in_elem.text)
                            out_frames = int(out_elem.text)

                            source_in = self.tc_handler.frames_to_timecode(in_frames)
                            source_out = self.tc_handler.frames_to_timecode(out_frames)
                            duration_frames = out_frames - in_frames
                            source_duration = self.tc_handler.frames_to_timecode(duration_frames)

                            if filename not in clip_timecodes:
                                clip_timecodes[filename] = []

                            clip_timecodes[filename].append({
                                'source_in': source_in,
                                'source_out': source_out,
                                'source_duration': source_duration
                            })
                        except (ValueError, TypeError):
                            continue

        return clip_timecodes

    def extract_metadata_row(self, file_path: str, framerate: int = 24,
                              clip_timecodes: Dict = None) -> Dict:
        """Extract metadata from an audio file."""
        filename = os.path.basename(file_path)
        cue_name = filename

        timecode_info = ""
        if clip_timecodes and filename in clip_timecodes:
            timecode_entries = []
            for tc in clip_timecodes[filename]:
                timecode_entries.append(f"In: {tc['source_in']} Out: {tc['source_out']} Duration: {tc['source_duration']}")
            timecode_info = " | ".join(timecode_entries)

        row = {
            'CUE NAME': cue_name,
            'FILE NAME': filename,
            'DURATION': '',
            'SOURCE TIMECODES': timecode_info,
            'ALBUM': '',
            'ALBUM/PERFORMER': '',
            'TRACK #': '',
            'GROUPING': '',
            'PERFORMER': '',
            'COMPOSER': '',
            'GENRE': '',
            'YEAR': '',
            'COMMENT': ''
        }

        try:
            media_info = MediaInfo.parse(file_path)
            tc_handler = TimecodeHandler(framerate)

            for track in media_info.tracks:
                if track.track_type == "General":
                    duration_ms = getattr(track, 'duration', None)
                    if duration_ms:
                        row['DURATION'] = tc_handler.ms_to_timecode(duration_ms)
                    row['ALBUM'] = getattr(track, 'album', '') or ''
                    row['ALBUM/PERFORMER'] = getattr(track, 'album_performer', '') or ''
                    row['TRACK #'] = getattr(track, 'track_name_position', '') or ''
                    row['GROUPING'] = getattr(track, 'grouping', '') or ''
                    row['PERFORMER'] = getattr(track, 'performer', '') or ''
                    row['COMPOSER'] = getattr(track, 'composer', '') or ''
                    row['GENRE'] = getattr(track, 'genre', '') or ''
                    row['YEAR'] = getattr(track, 'recorded_date', '') or ''
                    row['COMMENT'] = getattr(track, 'comment', '') or ''
        except Exception as e:
            logger.error(f"Error extracting metadata from {file_path}: {e}")

        return row

    def extract_project_info_from_xml(self, root, framerate: int) -> Dict[str, str]:
        """Extract project information from parsed XML root for template header."""
        project_info = {
            'content_producer': '',
            'runtime': '',
            'project_title': '',
            'version': '',
            'featurette_title': '',
            'date': datetime.now().strftime("%Y-%m-%d")
        }

        sequence = root.find('.//sequence')
        if sequence is not None:
            name_elem = sequence.find('name')
            if name_elem is not None and name_elem.text:
                project_info['project_title'] = name_elem.text
                project_info['featurette_title'] = name_elem.text

            duration_elem = sequence.find('duration')
            if duration_elem is not None and duration_elem.text:
                frames = int(duration_elem.text)
                tc_handler = TimecodeHandler(framerate)
                runtime_timecode = tc_handler.frames_to_timecode(frames)
                project_info['runtime'] = runtime_timecode

        return project_info

    def populate_excel_template(self, template_path: str, output_dir: str,
                                 audio_data: List[Dict], project_info: Dict) -> str:
        """Populate Excel template with extracted data."""
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")

        output_filename = f"CueSheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        shutil.copy2(template_path, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Populate header information
        if project_info['content_producer']:
            ws['D2'] = project_info['content_producer']
        if project_info['runtime']:
            ws['H2'] = project_info['runtime']
        if project_info['project_title']:
            ws['D3'] = project_info['project_title']
        if project_info['version']:
            ws['H3'] = project_info['version']
        if project_info['featurette_title']:
            ws['D4'] = project_info['featurette_title']
        ws['H4'] = project_info['date']

        # Populate cue data starting from row 7
        cue_number = 0
        for row_data in audio_data:
            if row_data.get('_is_separator'):
                continue
            cue_number += 1
            i = cue_number - 1
            row_num = 7 + i

            ws[f'A{row_num}'] = f"{cue_number}.0"

            timecode_info = row_data.get('SOURCE TIMECODES', '')
            tc_in, tc_out = self._parse_timecode_info(timecode_info)

            ws[f'B{row_num}'] = tc_in
            if 'Out:' in timecode_info:
                ws[f'C{row_num}'] = tc_out
            else:
                ws[f'C{row_num}'] = row_data.get('DURATION', '00:00:00')

            ws[f'D{row_num}'] = row_data.get('CUE NAME', '')
            ws[f'E{row_num}'] = row_data.get('DURATION', '00:00:00')

            filename = row_data.get('FILE NAME', '').lower()
            if 'instrumental' in filename or 'inst' in filename:
                usage = 'BI'
            elif 'vocal' in filename or 'voice' in filename:
                usage = 'BV'
            else:
                usage = 'BV'
            ws[f'F{row_num}'] = usage

            ws[f'G{row_num}'] = 'Library'
            ws[f'H{row_num}'] = 'In Context'

            composer = row_data.get('COMPOSER', '')
            ws[f'I{row_num}'] = composer if composer else 'Unknown/Unknown'

            ws[f'J{row_num}'] = row_data.get('ALBUM', '') or 'Unknown'
            ws[f'K{row_num}'] = row_data.get('ALBUM/PERFORMER', '') or 'Library'

        wb.save(output_path)
        return output_path

    def process(self, xml_path: str, template_path: Optional[str], output_dir: str,
                filter_keywords: List[str] = None,
                progress_callback: Callable = None) -> Dict:
        """Complete music cue processing pipeline.

        Args:
            xml_path: Path to Premiere XML file
            template_path: Optional path to Excel template
            output_dir: Directory for output files
            filter_keywords: Ignored (kept for backward compatibility).
            progress_callback: Optional progress callback function
        """
        results = {
            'status': 'success',
            'message': '',
            'files': {},
            'cue_count': 0
        }

        try:
            if progress_callback:
                progress_callback(0.1, "Parsing XML file...")

            root = self._parse_xml(xml_path)
            framerate = self._extract_framerate_from_root(root)

            if progress_callback:
                progress_callback(0.15, "Extracting project information...")

            project_info = self.extract_project_info_from_xml(root, framerate)

            if progress_callback:
                progress_callback(0.2, "Extracting audio paths...")

            audio_paths = self.extract_audio_paths_from_xml(root)
            if not audio_paths:
                results['status'] = 'warning'
                results['message'] = 'No audio files found in the XML'
                return results

            if progress_callback:
                progress_callback(0.3, "Extracting clip timecodes...")

            clip_timecodes = self.extract_clip_timecodes_from_xml(root, framerate)

            if progress_callback:
                progress_callback(0.4, "Processing audio metadata...")

            processed_filenames = set()
            audio_data = []
            total_files = len(audio_paths)

            for i, file_path in enumerate(audio_paths):
                if progress_callback:
                    progress = 0.4 + (0.4 * i / total_files)
                    progress_callback(progress, f"Processing file {i + 1}/{total_files}")

                if not os.path.exists(file_path):
                    logger.warning(f"File not found: {file_path}")
                    continue

                filename = os.path.basename(file_path)
                if filename in processed_filenames:
                    continue

                row_data = self.extract_metadata_row(file_path, framerate, clip_timecodes)
                audio_data.append(row_data)
                processed_filenames.add(filename)

            # Sort by metadata richness (rich first), then alphabetically
            audio_data.sort(key=lambda r: (-self._metadata_richness_score(r), r['CUE NAME'].lower()))

            # Insert separator between rich and poor metadata rows
            audio_data = self._insert_separator(audio_data)

            if progress_callback:
                progress_callback(0.9, "Generating Excel output...")

            if template_path and os.path.exists(template_path):
                output_path = self.populate_excel_template(template_path, output_dir, audio_data, project_info)
            else:
                output_path = self._create_basic_cue_sheet(output_dir, audio_data, project_info)

            if progress_callback:
                progress_callback(1.0, "Processing complete!")

            # Count excludes separator row
            real_count = sum(1 for r in audio_data if not r.get('_is_separator'))
            results['files']['excel'] = output_path
            results['cue_count'] = real_count
            results['message'] = f"Successfully processed {real_count} music cues"

            return results

        except Exception as e:
            logger.error(f"Music cue processing failed: {e}")
            results['status'] = 'error'
            results['message'] = str(e)
            return results

    @staticmethod
    def _insert_separator(audio_data: List[Dict]) -> List[Dict]:
        """Insert a separator row between metadata-rich and metadata-poor files."""
        if not audio_data:
            return audio_data

        # Find boundary between rich (score > 0) and poor (score == 0)
        last_rich_idx = -1
        for i, row in enumerate(audio_data):
            score = sum(1 for f in _RICHNESS_FIELDS if row.get(f, '').strip())
            if score > 0:
                last_rich_idx = i

        # Only insert separator if there are both rich and poor rows
        first_poor_exists = any(
            sum(1 for f in _RICHNESS_FIELDS if r.get(f, '').strip()) == 0
            for r in audio_data
        )

        if last_rich_idx >= 0 and first_poor_exists:
            separator = {
                'CUE NAME': '--- Files Without Metadata ---',
                'FILE NAME': '',
                'DURATION': '',
                'SOURCE TIMECODES': '',
                'ALBUM': '',
                'ALBUM/PERFORMER': '',
                'TRACK #': '',
                'GROUPING': '',
                'PERFORMER': '',
                'COMPOSER': '',
                'GENRE': '',
                'YEAR': '',
                'COMMENT': '',
                '_is_separator': True,
            }
            audio_data.insert(last_rich_idx + 1, separator)

        return audio_data

    def _create_basic_cue_sheet(self, output_dir: str, audio_data: List[Dict],
                                 project_info: Dict) -> str:
        """Create a professionally formatted cue sheet without a template."""
        output_filename = f"CueSheet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(output_dir, output_filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Music Cues"

        # -- Style definitions --
        header_font = Font(name='Calibri', size=11, bold=True)
        header_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(bottom=Side(style='thin', color='555555'))

        data_font = Font(name='Calibri', size=11)
        separator_font = Font(name='Calibri', size=11, italic=True, color='888888')
        separator_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        even_fill = PatternFill(fill_type=None)
        odd_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

        cell_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0'),
        )

        # -- Headers --
        headers = ['Cue #', 'TC In', 'TC Out', 'Cue Name', 'Duration', 'Use',
                   'Source', 'Context', 'Composer', 'Publisher', 'Library']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        ws.row_dimensions[1].height = 25

        # -- Data rows --
        cue_number = 0
        data_row_idx = 0  # track for zebra striping

        for row_data in audio_data:
            row_num = data_row_idx + 2  # starts at row 2
            is_separator = row_data.get('_is_separator', False)

            if is_separator:
                # Separator row spans the cue name column
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = separator_font
                    cell.fill = separator_fill
                    cell.border = cell_border
                ws.cell(row=row_num, column=4, value=row_data['CUE NAME'])
                ws.row_dimensions[row_num].height = 20
                data_row_idx += 1
                continue

            cue_number += 1
            timecode_info = row_data.get('SOURCE TIMECODES', '')
            tc_in, tc_out = self._parse_timecode_info(timecode_info)
            if tc_out == "00:00:00" and 'Out:' not in timecode_info:
                tc_out = row_data.get('DURATION', '00:00:00')

            filename = row_data.get('FILE NAME', '').lower()
            usage = 'BI' if 'instrumental' in filename or 'inst' in filename else 'BV'

            values = [
                f"{cue_number}.0",
                tc_in,
                tc_out,
                row_data.get('CUE NAME', ''),
                row_data.get('DURATION', ''),
                usage,
                'Library',
                'In Context',
                row_data.get('COMPOSER', 'Unknown'),
                row_data.get('ALBUM', 'Unknown'),
                row_data.get('ALBUM/PERFORMER', 'Library'),
            ]

            # Zebra striping (alternates on non-separator data rows)
            stripe = odd_fill if (cue_number % 2 == 0) else even_fill

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = data_font
                cell.border = cell_border
                if stripe.fill_type:
                    cell.fill = stripe

            ws.row_dimensions[row_num].height = 16
            data_row_idx += 1

        # -- Auto-fit column widths --
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_len = len(headers[col - 1])
            for row in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        # -- Frozen header row --
        ws.freeze_panes = 'A2'

        # -- Print setup --
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
            fitToPage=True
        )
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = '1:1'

        wb.save(output_path)
        return output_path
