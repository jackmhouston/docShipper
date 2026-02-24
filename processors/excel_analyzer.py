"""
Advanced Excel Template Analyzer for DocShipper
Uses pandas + fuzzywuzzy for superior pattern matching and table detection
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz
import logging
from typing import Dict, List
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AdvancedExcelAnalyzer:
    """
    Advanced Excel analyzer using pandas and fuzzywuzzy for superior
    pattern matching and table structure detection
    """

    def __init__(self):
        # Comprehensive industry terminology patterns
        self.field_patterns = {
            'clip_name': [
                'source file name', 'source filename', 'clip name', 'clip_name',
                'media file', 'video file', 'asset name', 'file name', 'filename',
                'source name', 'clip file', 'media name', 'source clip',
                'asset', 'clip title', 'clip', 'source', 'name'
            ],
            'src_start': [
                'time code in', 'timecode in', 'tc in', 'source timecode in',
                'src start', 'source start', 'source in', 'clip in',
                'in point', 'start time', 'timecode start', 'source tc in',
                'clip start', 'media in', 'start tc', 'in', 'start'
            ],
            'src_end': [
                'time code out', 'timecode out', 'tc out', 'source timecode out',
                'src end', 'source end', 'source out', 'clip out',
                'out point', 'end time', 'timecode end', 'source tc out',
                'clip end', 'media out', 'end tc', 'out', 'end'
            ],
            'rec_start': [
                'record in', 'rec start', 'timeline in', 'program in',
                'sequence in', 'edit in', 'promo start', 'show in',
                'timeline start', 'record start', 'timeline tc in', 'record tc in'
            ],
            'rec_end': [
                'record out', 'rec end', 'timeline out', 'program out',
                'sequence out', 'edit out', 'promo end', 'show out',
                'timeline end', 'record end', 'timeline tc out', 'record tc out'
            ],
            'duration': [
                'duration', 'length', 'runtime', 'run time', 'clip length',
                'media duration', 'total time', 'time', 'clip duration'
            ],
            'screenshot': [
                'screen shot', 'screenshot', 'thumbnail', 'image', 'preview',
                'frame grab', 'still', 'capture', 'thumb', 'keyframe',
                'reference image', 'preview image', 'picture', 'frame'
            ]
        }

        self.min_confidence = 85
        self.exact_match_bonus = 10

    def analyze_template(self, file_path: str) -> Dict[str, str]:
        """
        Analyze Excel template with advanced pattern matching.
        Returns mapping of field names to Excel cell references.
        """
        try:
            logger.info(f"Analyzing template: {file_path}")

            df = pd.read_excel(file_path, sheet_name=0, header=None, nrows=30)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            logger.info(f"Template dimensions: {df.shape[0]} rows x {df.shape[1]} columns")

            headers = self._find_all_headers(df, ws)
            logger.info(f"Found {len(headers)} potential headers")

            mappings = self._match_headers_to_fields(headers, ws)
            final_mappings = self._validate_mappings(mappings, ws)

            logger.info(f"Final mappings: {final_mappings}")
            return final_mappings

        except Exception as e:
            logger.error(f"Error analyzing template: {e}")
            return {}

    def _find_all_headers(self, df: pd.DataFrame, ws) -> List[Dict]:
        """Find all potential header cells in the spreadsheet."""
        headers = []

        for row_idx in range(min(25, df.shape[0])):
            for col_idx in range(min(20, df.shape[1])):
                cell_value = df.iloc[row_idx, col_idx]

                if pd.notna(cell_value) and isinstance(cell_value, str):
                    cell_value = str(cell_value).strip()

                    if len(cell_value) < 2 or cell_value.isdigit():
                        continue

                    if re.search(r'[a-zA-Z]', cell_value):
                        col_letter = get_column_letter(col_idx + 1)
                        excel_row = row_idx + 1

                        headers.append({
                            'text': cell_value,
                            'row': excel_row,
                            'col': col_idx + 1,
                            'col_letter': col_letter,
                            'cell_ref': f"{col_letter}{excel_row}",
                            'normalized': self._normalize_text(cell_value)
                        })

        return headers

    def _normalize_text(self, text: str) -> str:
        """Normalize text for better matching."""
        normalized = text.lower().strip()

        for sep in ['_', '-', '.', '/', '\\', ':', ';', '(', ')', '[', ']', '\n', '\r']:
            normalized = normalized.replace(sep, ' ')

        normalized = ' '.join(normalized.split())
        return normalized

    def _match_headers_to_fields(self, headers: List[Dict], ws) -> Dict[str, Dict]:
        """Match headers to field types using advanced fuzzy matching."""
        mappings = {}

        for field, patterns in self.field_patterns.items():
            best_match = None
            best_score = 0

            for header in headers:
                header_text = header['normalized']

                for pattern in patterns:
                    score = fuzz.ratio(header_text, pattern)

                    if pattern in header_text or header_text in pattern:
                        score += self.exact_match_bonus

                    if fuzz.token_sort_ratio(header_text, pattern) > score:
                        score = fuzz.token_sort_ratio(header_text, pattern)

                    if score > best_score and score >= self.min_confidence:
                        best_score = score
                        best_match = {
                            'header': header,
                            'pattern': pattern,
                            'score': score,
                            'field': field
                        }

            if best_match:
                mappings[field] = best_match
                logger.info(f"Matched {field}: '{best_match['header']['text']}' "
                            f"at {best_match['header']['cell_ref']} "
                            f"(score: {best_score}, pattern: '{best_match['pattern']}')")

        return mappings

    def _validate_mappings(self, mappings: Dict[str, Dict], ws) -> Dict[str, str]:
        """Validate mappings and find appropriate data insertion points."""
        final_mappings = {}

        for field, match_info in mappings.items():
            header = match_info['header']

            if header['row'] == 13 and field in ['clip_name', 'src_start', 'src_end', 'screenshot']:
                data_row = 15
                logger.info(f"Using NGO template override for {field}: row 15")
            else:
                data_row = self._find_data_insertion_row(ws, header['col'], header['row'])

            cell_ref = f"{header['col_letter']}{data_row}"
            final_mappings[field] = cell_ref

            logger.info(f"Final mapping {field}: {cell_ref}")

        return final_mappings

    def _find_data_insertion_row(self, ws, col: int, header_row: int) -> int:
        """Find the appropriate row to start inserting data."""
        start_row = header_row + 1

        for row in range(start_row, start_row + 10):
            cell_value = ws.cell(row=row, column=col).value
            col_a_value = ws.cell(row=row, column=1).value

            if cell_value is None:
                if isinstance(col_a_value, int) or (isinstance(col_a_value, str) and str(col_a_value).strip().isdigit()):
                    return row

                if row > 1:
                    prev_col_a = ws.cell(row=row - 1, column=1).value
                    if isinstance(prev_col_a, str) and prev_col_a.lower().strip() == 'example':
                        return row

                return row

            if isinstance(cell_value, (int, str)) and str(cell_value).strip().isdigit():
                return row

        if header_row == 13:
            return 15

        return start_row

    def get_analysis_summary(self, mappings: Dict[str, str]) -> str:
        """Generate a human-readable summary of the analysis."""
        if not mappings:
            return "No field mappings detected. Please check your template format."

        summary = f"Successfully detected {len(mappings)} field mappings:\n\n"

        field_names = {
            'clip_name': 'Clip/Source File Name',
            'src_start': 'Source Start Time',
            'src_end': 'Source End Time',
            'rec_start': 'Record Start Time',
            'rec_end': 'Record End Time',
            'duration': 'Duration',
            'screenshot': 'Screenshot/Image'
        }

        for field, cell_ref in mappings.items():
            field_display = field_names.get(field, field.replace('_', ' ').title())
            summary += f"- {field_display}: Cell {cell_ref}\n"

        critical_fields = ['clip_name', 'src_start', 'screenshot']
        missing = [f for f in critical_fields if f not in mappings]

        if missing:
            summary += f"\nMissing critical fields: {', '.join(missing)}\n"
            summary += "These fields are recommended for complete shotlist generation.\n"

        return summary
