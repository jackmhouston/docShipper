"""
Video Processing Module for DocShipper
Handles EDL parsing, video analysis, screenshot generation, and Excel updates
"""

import subprocess
import json
import logging
import os
import re
import defusedxml.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Callable
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.timecode import TimecodeHandler
from utils import sanitize_filename

logger = logging.getLogger(__name__)


class EDLParser:
    """Enhanced EDL parser for various formats."""

    def __init__(self):
        self.clip_patterns = [
            r'^\* FROM CLIP NAME:\s*(.+)$',
            r'^\* CLIP NAME:\s*(.+)$',
            r'^\*\s*(.+\.mov)$',
            r'^\*\s*(.+\.mp4)$',
        ]

    def _is_valid_timecode(self, tc: str) -> bool:
        """Check if a string looks like a valid timecode (HH:MM:SS:FF)."""
        return TimecodeHandler.is_valid_timecode(tc)

    def parse(self, edl_path: str) -> List[Dict]:
        """Parse EDL file and extract shot information."""
        try:
            with open(edl_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()
        except UnicodeDecodeError:
            with open(edl_path, 'r', encoding='latin-1') as file:
                lines = file.readlines()

        edl_data = []
        current_event = None
        line_number = 0

        logger.info(f"Parsing EDL: {edl_path}")
        logger.info(f"Total lines in EDL: {len(lines)}")

        for line in lines:
            line_number += 1
            line = line.strip()

            if not line:
                continue

            # Check for edit events (lines starting with numbers)
            if line and line[0].isdigit():
                if current_event:
                    edl_data.append(current_event)
                current_event = None

                line_parts = line.split()
                if len(line_parts) >= 8:
                    track_type = line_parts[2]

                    # Only process VIDEO tracks
                    if track_type.upper().startswith('V'):
                        reel_name = line_parts[1]

                        # Skip BL (blank/black) events
                        if reel_name.upper() == 'BL':
                            logger.debug(f"Line {line_number}: Skipping blank (BL) event")
                            continue

                        src_start = line_parts[4]
                        src_end = line_parts[5]
                        rec_start = line_parts[6]
                        rec_end = line_parts[7]

                        if not (self._is_valid_timecode(src_start) and
                                self._is_valid_timecode(src_end) and
                                self._is_valid_timecode(rec_start) and
                                self._is_valid_timecode(rec_end)):
                            logger.warning(f"Line {line_number}: Invalid timecode format, skipping")
                            continue

                        current_event = {
                            'Event': line_parts[0],
                            'Reel': reel_name,
                            'Track': track_type,
                            'Transition': line_parts[3],
                            'Src Start': src_start,
                            'Src End': src_end,
                            'Rec Start': rec_start,
                            'Rec End': rec_end,
                            'Clip Name': None
                        }
                        logger.debug(f"Line {line_number}: Parsed video track event {line_parts[0]}")
                continue

            # Look for clip names
            for pattern in self.clip_patterns:
                match = re.match(pattern, line, re.IGNORECASE)
                if match:
                    clip_name = match.group(1).strip()
                    if current_event:
                        current_event['Clip Name'] = clip_name
                    break

        if current_event:
            edl_data.append(current_event)

        logger.info(f"Parsed {len(edl_data)} events from EDL")
        return edl_data


class XMLParser:
    """Parser for Premiere Pro Final Cut Pro XML exports.

    XML provides exact frame numbers which bypass timecode conversion issues entirely.
    This is more accurate than EDL parsing for NTSC frame rates.
    """

    def __init__(self):
        self.frame_rate = None
        self.ntsc = False

    def parse(self, xml_path: str) -> List[Dict]:
        """Parse Premiere Pro XML file and extract shot information.

        Returns data in the same format as EDLParser for compatibility.
        """
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
        except ET.ParseError as e:
            logger.error(f"Failed to parse XML file {xml_path}: {e}")
            return []

        xml_data = []

        # Find the sequence element
        sequence = root.find('.//sequence')
        if sequence is None:
            logger.error("No sequence found in XML file")
            return []

        # Get sequence frame rate
        rate_elem = sequence.find('.//rate')
        if rate_elem is not None:
            timebase = rate_elem.find('timebase')
            ntsc = rate_elem.find('ntsc')
            if timebase is not None:
                self.frame_rate = float(timebase.text)
                if ntsc is not None and ntsc.text.upper() == 'TRUE':
                    self.ntsc = True
                    # Adjust for NTSC rates
                    if self.frame_rate == 24:
                        self.frame_rate = 23.976
                    elif self.frame_rate == 30:
                        self.frame_rate = 29.97
                    elif self.frame_rate == 60:
                        self.frame_rate = 59.94

        logger.info(f"XML sequence frame rate: {self.frame_rate} fps (NTSC: {self.ntsc})")

        # Find all video tracks
        video_tracks = sequence.findall('.//video/track')
        event_number = 0

        for track in video_tracks:
            clipitems = track.findall('clipitem')

            for clipitem in clipitems:
                # Skip disabled clips
                enabled = clipitem.find('enabled')
                if enabled is not None and enabled.text.upper() == 'FALSE':
                    continue

                # Get clip name
                name_elem = clipitem.find('name')
                clip_name = name_elem.text if name_elem is not None else None

                # Skip black/blank clips
                if clip_name and clip_name.lower() in ['black video', 'black', 'blank']:
                    continue

                # Get frame positions (these are exact - no conversion needed!)
                start_elem = clipitem.find('start')
                end_elem = clipitem.find('end')
                in_elem = clipitem.find('in')
                out_elem = clipitem.find('out')

                # Skip if missing critical data
                if start_elem is None or end_elem is None:
                    continue

                # Handle -1 values (which mean "use in/out points")
                start_frame = int(start_elem.text) if start_elem.text != '-1' else 0
                end_frame = int(end_elem.text) if end_elem.text != '-1' else 0

                # Get source in/out points
                in_frame = int(in_elem.text) if in_elem is not None and in_elem.text else 0
                out_frame = int(out_elem.text) if out_elem is not None and out_elem.text else 0

                # Skip if this is actually a gap (start == end or negative duration)
                if end_frame <= start_frame:
                    continue

                event_number += 1

                # Convert frames to timecode format for compatibility with existing code
                tc_handler = TimecodeHandler(self.frame_rate if self.frame_rate else 24.0)

                # Record timecodes (position in timeline)
                rec_start_tc = tc_handler.frames_to_timecode(start_frame)
                rec_end_tc = tc_handler.frames_to_timecode(end_frame)

                # Source timecodes (position in source clip)
                src_start_tc = tc_handler.frames_to_timecode(in_frame)
                src_end_tc = tc_handler.frames_to_timecode(out_frame)

                event = {
                    'Event': str(event_number).zfill(3),
                    'Reel': 'AX',
                    'Track': 'V',
                    'Transition': 'C',
                    'Src Start': src_start_tc,
                    'Src End': src_end_tc,
                    'Rec Start': rec_start_tc,
                    'Rec End': rec_end_tc,
                    'Clip Name': clip_name,
                    # Store exact frame numbers for precise seeking
                    '_start_frame': start_frame,
                    '_end_frame': end_frame,
                    '_in_frame': in_frame,
                    '_out_frame': out_frame,
                }

                xml_data.append(event)
                logger.debug(f"Parsed clip {event_number}: {clip_name} at frame {start_frame}-{end_frame}")

        logger.info(f"Parsed {len(xml_data)} clips from XML")
        return xml_data

    def get_frame_rate(self) -> Optional[float]:
        """Return detected XML frame rate."""
        return self.frame_rate

class OTIOParser:
    """Parser for OpenTimelineIO JSON exports (timeline representation)."""

    def __init__(self):
        self.frame_rate = 24.0

    def parse(self, otio_path: str) -> List[Dict]:
        """Parse an OTIO JSON timeline and extract clip-level data."""
        try:
            with open(otio_path, 'r', encoding='utf-8') as stream:
                data = json.load(stream)
        except (json.JSONDecodeError, OSError) as e:
            logger.error(f"Failed to parse OTIO file {otio_path}: {e}")
            return []

        self.frame_rate = self._extract_frame_rate(data) or 24.0
        track = self._find_primary_video_track(data)
        if track is None:
            logger.error("No target video track found in OTIO file")
            return []

        clips = self._collect_clips(track)
        if not clips:
            logger.warning("OTIO track contains no clip items")
            return []

        tc_handler = TimecodeHandler(self.frame_rate)
        otio_data = []
        for idx, clip in enumerate(clips, start=1):
            source_range = clip.get('source_range')
            if not source_range:
                continue

            start_val = self._safe_value(source_range, 'start_time', 'value')
            duration_val = self._safe_value(source_range, 'duration', 'value')
            if duration_val is None or start_val is None or duration_val <= 0:
                continue

            start_frame = int(round(start_val))
            duration_frames = int(round(duration_val))
            end_frame = start_frame + duration_frames

            rec_start_tc = tc_handler.frames_to_timecode(start_frame)
            rec_end_tc = tc_handler.frames_to_timecode(end_frame)
            duration_tc = tc_handler.frames_to_timecode(duration_frames)

            clip_name = clip.get('name') or 'Untitled'

            otio_data.append({
                'Event': str(idx).zfill(3),
                'Reel': 'AX',
                'Track': 'V',
                'Transition': 'C',
                'Src Start': rec_start_tc,
                'Src End': rec_end_tc,
                'Rec Start': rec_start_tc,
                'Rec End': rec_end_tc,
                'Duration': duration_tc,
                'Clip Name': clip_name,
                '_start_frame': start_frame,
                '_end_frame': end_frame,
                '_in_frame': start_frame,
                '_out_frame': end_frame,
            })

        logger.info(f"Parsed {len(otio_data)} clips from OTIO")
        return otio_data

    def _extract_frame_rate(self, data: Dict) -> Optional[float]:
        tracks = data.get('tracks', {})
        metadata = tracks.get('metadata', {})
        premiere = metadata.get('PremierePro_OTIO', {})
        if premiere.get('VideoFrameRate'):
            return float(premiere['VideoFrameRate'])
        global_start = data.get('global_start_time', {})
        rate = global_start.get('rate')
        if rate:
            return float(rate)
        return None

    def _find_primary_video_track(self, data: Dict) -> Optional[Dict]:
        tracks = data.get('tracks', {}).get('children', [])
        for track in tracks:
            name = track.get('name', '').lower()
            if not track.get('enabled', False):
                continue
            if 'video' not in name:
                continue
            if any(self._is_clip_candidate(child) for child in track.get('children', [])):
                return track
        return None

    def _collect_clips(self, track: Dict) -> List[Dict]:
        clips = [
            child for child in track.get('children', [])
            if self._is_clip_candidate(child)
        ]
        return sorted(clips, key=lambda c: self._safe_value(c.get('source_range', {}), 'start_time', 'value') or 0)

    @staticmethod
    def _is_clip_candidate(node: Dict) -> bool:
        schema = node.get('OTIO_SCHEMA', '')
        return isinstance(schema, str) and schema.lower().startswith('clip')

    @staticmethod
    def _safe_value(node: Dict, *keys):
        current = node
        for key in keys:
            if not isinstance(current, dict):
                return None
            current = current.get(key)
        return current

    def get_frame_rate(self) -> float:
        return self.frame_rate

    def _frames_to_timecode(self, frames: int, nominal_rate: int) -> str:
        """Convert frame count to timecode string using nominal rate."""
        if frames < 0:
            frames = 0
        frame_remainder = frames % nominal_rate
        total_seconds = frames // nominal_rate
        seconds = total_seconds % 60
        total_minutes = total_seconds // 60
        minutes = total_minutes % 60
        hours = total_minutes // 60
        return f"{hours:02}:{minutes:02}:{seconds:02}:{frame_remainder:02}"

    def get_frame_rate(self) -> Optional[float]:
        """Return the detected frame rate from the XML."""
        return self.frame_rate


class VideoAnalyzer:
    """Video file analysis and frame rate detection."""

    def get_video_frame_rate(self, video_file: str) -> Optional[float]:
        """Get video frame rate using ffprobe."""
        command = [
            'ffprobe',
            '-v', 'error',
            '-select_streams', 'v:0',
            '-show_entries', 'stream=r_frame_rate,avg_frame_rate,time_base',
            '-of', 'json',
            video_file
        ]

        try:
            result = subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output = json.loads(result.stdout)

            r_frame_rate = output['streams'][0].get('r_frame_rate', '')
            if r_frame_rate:
                try:
                    num, den = map(int, r_frame_rate.split('/'))
                    if num and den:
                        frame_rate = num / den
                        if abs(frame_rate - 23.976) < 0.01:
                            return 23.976
                        if abs(frame_rate - 29.97) < 0.01:
                            return 29.97
                        return frame_rate
                except (ValueError, ZeroDivisionError):
                    pass

            avg_frame_rate = output['streams'][0].get('avg_frame_rate', '')
            if avg_frame_rate:
                try:
                    num, den = map(int, avg_frame_rate.split('/'))
                    if num and den:
                        return num / den
                except (ValueError, ZeroDivisionError):
                    pass

            logger.error(f'Failed to parse frame rate for {video_file}')
            return None

        except subprocess.CalledProcessError as e:
            logger.error(f'Failed to get frame rate for {video_file}: {e}')
            return None


class ScreenshotGenerator:
    """Advanced screenshot generation with retry logic."""

    def __init__(self, video_analyzer: VideoAnalyzer):
        self.video_analyzer = video_analyzer

    def capture_screenshot(self, video_file: str, timecode_start: str, timecode_end: str,
                           shot_number: int, clip_name: str, frame_rate: float,
                           screenshot_dir: str, start_frame: int = None,
                           end_frame: int = None,
                           width: int = 203, height: int = 120,
                           quality: int = 2) -> bool:
        """Capture screenshot at specific timecode with retry logic.

        Args:
            video_file: Path to video file
            timecode_start: Start timecode (used if start_frame not provided)
            timecode_end: End timecode (used if end_frame not provided)
            shot_number: Sequential shot number for naming
            clip_name: Clip name for filename
            frame_rate: Video frame rate
            screenshot_dir: Output directory
            start_frame: Exact start frame from XML (optional, more accurate)
            end_frame: Exact end frame from XML (optional, more accurate)
            width: Screenshot width in pixels
            height: Screenshot height in pixels
            quality: FFmpeg quality parameter (1-10, lower is better)
        """
        try:
            tc_handler = TimecodeHandler(frame_rate)

            # Use exact frame numbers if available (from XML), otherwise convert from timecode
            if start_frame is not None and end_frame is not None:
                start_seconds = tc_handler.frames_to_seconds(start_frame)
                end_seconds = tc_handler.frames_to_seconds(end_frame)
                logger.info(f"Shot {shot_number}: XML frame {start_frame} -> {start_seconds:.6f}s (frame-accurate)")
            else:
                start_seconds = tc_handler.timecode_to_seconds(timecode_start)
                end_seconds = tc_handler.timecode_to_seconds(timecode_end)
                # Log conversion details for debugging
                nominal = tc_handler.nominal_rate
                parts = timecode_start.split(':')
                h, m, s, f = map(int, parts)
                calculated_frames = h * 3600 * nominal + m * 60 * nominal + s * nominal + f
                logger.info(f"Shot {shot_number}: TC {timecode_start} -> frame {calculated_frames} -> {start_seconds:.6f}s")

            clip_duration = end_seconds - start_seconds

            one_frame = 1.0 / frame_rate
            max_offset = clip_duration * 0.1
            safe_offset = min(one_frame, max_offset) if clip_duration > 0 else 0

            time_in_seconds = start_seconds + safe_offset

            if clip_name is None or clip_name.strip() == '':
                clean_clip_name = f"Shot{str(shot_number).zfill(3)}"
            else:
                clean_clip_name = f"Shot{str(shot_number).zfill(3)}_{clip_name}"
                clean_clip_name = sanitize_filename(clean_clip_name)
                clean_clip_name = clean_clip_name[:200]

            output_file = os.path.join(screenshot_dir, f'{clean_clip_name}.png')
            os.makedirs(screenshot_dir, exist_ok=True)

            # Use two-pass seeking for frame-accurate capture on H.264/compressed video:
            # 1. Fast seek to 2 seconds before target (keyframe seeking)
            # 2. Precise seek from there to exact position (frame-accurate)
            pre_seek = max(0, time_in_seconds - 2.0)
            precise_offset = time_in_seconds - pre_seek

            scale_filter = f'scale={width}:{height}:force_original_aspect_ratio=decrease,pad={width}:{height}:(ow-iw)/2:(oh-ih)/2'

            # Primary method: Two-pass seeking (most accurate for H.264)
            command = [
                'ffmpeg',
                '-y',
                '-ss', f"{pre_seek:.6f}",      # Fast seek to near the target
                '-i', video_file,
                '-ss', f"{precise_offset:.6f}", # Precise seek from there
                '-frames:v', '1',
                '-q:v', str(quality),
                '-vf', scale_filter,
                output_file
            ]

            logger.info(f"Processing shot {shot_number}: {clean_clip_name} at {time_in_seconds:.6f}s")
            result = subprocess.run(command, capture_output=True, text=True)

            if result.returncode != 0 or not os.path.exists(output_file) or os.path.getsize(output_file) == 0:
                # Fallback 1: Single pass with -ss after -i (slower but frame-accurate)
                logger.warning(f"Two-pass method failed for shot {shot_number}, trying frame-accurate fallback")
                command = [
                    'ffmpeg',
                    '-y',
                    '-i', video_file,
                    '-ss', f"{time_in_seconds:.6f}",
                    '-frames:v', '1',
                    '-q:v', str(quality),
                    '-vf', scale_filter,
                    output_file
                ]
                result = subprocess.run(command, capture_output=True, text=True)

            if result.returncode != 0 or not os.path.exists(output_file) or os.path.getsize(output_file) == 0:
                # Fallback 2: Original single-pass fast seek
                logger.warning(f"Frame-accurate method failed for shot {shot_number}, trying fast seek")
                command = [
                    'ffmpeg',
                    '-y',
                    '-ss', f"{time_in_seconds:.6f}",
                    '-i', video_file,
                    '-frames:v', '1',
                    '-q:v', str(quality),
                    '-vf', scale_filter,
                    output_file
                ]
                result = subprocess.run(command, capture_output=True, text=True)

            if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                logger.info(f"Successfully created screenshot: {output_file}")
                return True
            else:
                logger.error(f"Screenshot not created or empty for shot {shot_number}")
                return False

        except Exception as e:
            logger.error(f"Error capturing screenshot for shot {shot_number}: {e}")
            return False


class ExcelUpdater:
    """Excel template updating with image insertion."""

    def update_template(self, template_path: Optional[str], edl_data: List[Dict],
                        mappings: Dict[str, str], screenshot_dir: str,
                        output_path: str) -> str:
        """Update Excel template with EDL data and screenshots."""
        try:
            if template_path and os.path.exists(template_path):
                try:
                    wb = openpyxl.load_workbook(template_path)
                    ws = wb.active
                    logger.info(f"Successfully loaded template: {template_path}")
                except Exception as e:
                    logger.error(f"Failed to load template {template_path}: {e}")
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Shotlist"
                    self._add_headers(ws, mappings)
            else:
                logger.info("No template provided, creating new workbook")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Shotlist"
                self._add_headers(ws, mappings)

            field_mapping = {
                'clip_name': 'Clip Name',
                'src_start': 'Src Start',
                'src_end': 'Src End',
                'rec_start': 'Rec Start',
                'rec_end': 'Rec End',
                'duration': 'Duration'
            }

            for field, cell_ref in mappings.items():
                if field == 'screenshot':
                    continue

                edl_field = field_mapping.get(field)
                if not edl_field:
                    continue

                try:
                    col_letter = ''.join(filter(str.isalpha, cell_ref))
                    start_row = int(''.join(filter(str.isdigit, cell_ref)))

                    for i, shot in enumerate(edl_data):
                        value = shot.get(edl_field, '')
                        ws[f'{col_letter}{start_row + i}'] = value

                    logger.info(f"Successfully mapped {field} to column {col_letter}")

                except Exception as e:
                    logger.error(f"Error mapping {field} to cell {cell_ref}: {e}")
                    continue

            if 'screenshot' in mappings:
                self._insert_screenshots(ws, mappings['screenshot'], edl_data, screenshot_dir)

            self._apply_formatting(ws, mappings)

            wb.save(output_path)
            logger.info(f"Excel file saved as: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error updating Excel template: {e}")
            raise

    def _add_headers(self, ws, mappings: Dict[str, str]):
        """Add headers to new workbook."""
        header_names = {
            'clip_name': 'Clip Name',
            'src_start': 'Source Start',
            'src_end': 'Source End',
            'rec_start': 'Record Start',
            'rec_end': 'Record End',
            'duration': 'Duration',
            'screenshot': 'Screenshot'
        }

        for field, cell_ref in mappings.items():
            try:
                col_letter = ''.join(filter(str.isalpha, cell_ref))
                row_num = int(''.join(filter(str.isdigit, cell_ref)))
                header_row = row_num - 1

                header_name = header_names.get(field, field.replace('_', ' ').title())
                ws[f'{col_letter}{header_row}'] = header_name

            except Exception as e:
                logger.error(f"Error adding header for {field}: {e}")

    def _insert_screenshots(self, ws, screenshot_cell: str, edl_data: List[Dict], screenshot_dir: str,
                             img_width: int = 203, img_height: int = 120):
        """Insert screenshots into Excel."""
        try:
            col_letter = ''.join(filter(str.isalpha, screenshot_cell))
            start_row = int(''.join(filter(str.isdigit, screenshot_cell)))

            for i, shot in enumerate(edl_data):
                clip_name = shot.get('Clip Name', '') or ''
                if clip_name:
                    clean_clip_name = f"Shot{str(i + 1).zfill(3)}_{clip_name}"
                    clean_clip_name = sanitize_filename(clean_clip_name)
                    screenshot_name = f"{clean_clip_name}.png"
                else:
                    screenshot_name = f"Shot{str(i + 1).zfill(3)}.png"

                screenshot_path = os.path.join(screenshot_dir, screenshot_name)

                if not os.path.exists(screenshot_path):
                    fallback_name = f"Shot{str(i + 1).zfill(3)}.png"
                    fallback_path = os.path.join(screenshot_dir, fallback_name)
                    if os.path.exists(fallback_path):
                        screenshot_path = fallback_path

                if os.path.exists(screenshot_path):
                    try:
                        img = Image(screenshot_path)
                        img.width = img_width
                        img.height = img_height
                        cell_position = f'{col_letter}{start_row + i}'
                        ws.add_image(img, cell_position)
                        logger.info(f"Added screenshot to {cell_position}")
                    except Exception as e:
                        logger.error(f"Error adding screenshot {screenshot_path}: {e}")
                else:
                    logger.warning(f"Screenshot not found: {screenshot_path}")

            logger.info(f"Screenshots insertion completed for column {col_letter}")

        except Exception as e:
            logger.error(f"Error inserting screenshots: {e}")

    def _apply_formatting(self, ws, mappings: Dict[str, str]):
        """Apply production-ready formatting to the worksheet.

        Adds header styling, zebra striping, auto-fit column widths,
        data-type-specific alignment, borders, and print setup.
        """
        try:
            if ws.max_row is None or ws.max_row < 1:
                logger.info("Worksheet is empty, skipping formatting")
                return

            # -- Determine header row --
            # Headers are written one row above the data start row.
            # Find the minimum data row from mappings to derive header row.
            header_row = 1
            data_start_row = 2
            for field, cell_ref in mappings.items():
                try:
                    row_num = int(''.join(filter(str.isdigit, cell_ref)))
                    if row_num > 0:
                        data_start_row = row_num
                        header_row = row_num - 1
                        break
                except (ValueError, TypeError):
                    continue

            # -- Identify columns by field type for alignment --
            timecode_fields = {'src_start', 'src_end', 'rec_start', 'rec_end'}
            right_align_fields = timecode_fields | {'duration'}
            screenshot_cols = set()
            clip_name_cols = set()
            right_align_cols = set()

            for field, cell_ref in mappings.items():
                try:
                    col_letter = ''.join(filter(str.isalpha, cell_ref))
                    col_idx = openpyxl.utils.column_index_from_string(col_letter)
                except (ValueError, TypeError):
                    continue

                if field == 'screenshot':
                    screenshot_cols.add(col_idx)
                elif field == 'clip_name':
                    clip_name_cols.add(col_idx)
                if field in right_align_fields:
                    right_align_cols.add(col_idx)

            # -- Style definitions --
            default_font = Font(name='Calibri', size=11)
            header_font = Font(name='Calibri', size=11, bold=True)
            header_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            header_border = Border(
                bottom=Side(style='thin', color='555555')
            )

            even_fill = PatternFill(fill_type=None)  # White (default)
            odd_fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

            cell_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0'),
            )

            wrap_alignment = Alignment(vertical='top', wrap_text=True)
            right_alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            clip_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            max_col = ws.max_column or 1
            max_row = ws.max_row or 1

            # -- 1. Header row formatting --
            if header_row >= 1:
                ws.row_dimensions[header_row].height = 25
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = header_border

            # -- 2. Freeze pane below header row --
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

            # -- 3. Data row formatting --
            for row in range(data_start_row, max_row + 1):
                row_offset = row - data_start_row
                is_odd_data_row = (row_offset % 2 == 1)
                fill = odd_fill if is_odd_data_row else even_fill

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = default_font
                    cell.border = cell_border

                    # Apply zebra striping
                    if is_odd_data_row:
                        cell.fill = fill

                    # Data-type-specific alignment
                    if col in screenshot_cols:
                        cell.alignment = center_alignment
                    elif col in clip_name_cols:
                        cell.alignment = clip_alignment
                    elif col in right_align_cols:
                        cell.alignment = right_alignment
                    else:
                        cell.alignment = wrap_alignment

                # Increase row height for screenshot rows
                if screenshot_cols:
                    ws.row_dimensions[row].height = 100

            # -- 4. Auto-fit column widths --
            min_width = 10
            max_width = 50
            screenshot_width = 25

            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)

                if col in screenshot_cols:
                    ws.column_dimensions[col_letter].width = screenshot_width
                    continue

                longest = min_width
                for row in range(header_row, max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > longest:
                            longest = cell_len

                # Add padding and clamp
                width = min(longest + 2, max_width)
                width = max(width, min_width)
                ws.column_dimensions[col_letter].width = width

            # -- 5. Print setup --
            ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
                fitToPage=True
            )
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # 0 = as many pages as needed vertically
            ws.print_title_rows = f'{header_row}:{header_row}'

            logger.info("Applied professional formatting to worksheet")

        except Exception as e:
            logger.error(f"Error applying formatting: {e}")


class VideoProcessor:
    """Main video processing orchestrator."""

    def __init__(self):
        self.edl_parser = EDLParser()
        self.xml_parser = XMLParser()
        self.otio_parser = OTIOParser()
        self.video_analyzer = VideoAnalyzer()
        self.screenshot_generator = ScreenshotGenerator(self.video_analyzer)
        self.excel_updater = ExcelUpdater()

    def _detect_file_type(self, file_path: str) -> str:
        """Detect whether file is EDL or XML based on extension and content."""
        if not file_path:
            return 'unknown'

        ext = Path(file_path).suffix.lower()
        if ext == '.xml':
            return 'xml'
        elif ext == '.edl':
            return 'edl'
        elif ext == '.otio':
            return 'otio'

        # Try to detect by content
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline().strip()
                if first_line.startswith('<?xml') or first_line.startswith('<xmeml'):
                    return 'xml'
                elif 'TITLE:' in first_line or first_line[0].isdigit():
                    return 'edl'
        except Exception:
            pass

        return 'edl'  # Default to EDL

    def process(self, edl_path: str, video_path: str, template_path: str,
                mappings: Dict[str, str], output_dir: str,
                progress_callback: Callable = None,
                edl_data_override: Optional[List[Dict]] = None,
                disable_screenshots: bool = False,
                show_name: Optional[str] = None,
                screenshot_width: int = 203,
                screenshot_height: int = 120,
                screenshot_quality: int = 2) -> Dict[str, str]:
        """Complete processing pipeline.

        Supports both EDL and XML file formats. XML files provide more accurate
        frame positioning as they contain exact frame numbers.
        """

        results = {
            'status': 'success',
            'message': '',
            'files': {}
        }

        try:
            if progress_callback:
                progress_callback(0.1, "Preparing edit list data...")

            if edl_data_override is not None:
                edl_data = edl_data_override
                logger.info(f"Using EDL data override with {len(edl_data)} rows")
            else:
                # Auto-detect file type and parse accordingly
                file_type = self._detect_file_type(edl_path)
                logger.info(f"Detected file type: {file_type}")

                if file_type == 'xml':
                    edl_data = self.xml_parser.parse(edl_path)
                    if not edl_data:
                        raise Exception("No data found in XML file")
                    xml_fps = self.xml_parser.get_frame_rate()
                    logger.info(f"Parsed {len(edl_data)} clips from XML at {xml_fps} fps (frame-accurate mode)")
                    logger.info("Using exact frame numbers from XML - this provides maximum accuracy")
                elif file_type == 'otio':
                    edl_data = self.otio_parser.parse(edl_path)
                    if not edl_data:
                        raise Exception("No data found in OTIO file")
                    otio_fps = self.otio_parser.get_frame_rate()
                    logger.info(f"Parsed {len(edl_data)} clips from OTIO at {otio_fps} fps")
                else:
                    edl_data = self.edl_parser.parse(edl_path)
                    if not edl_data:
                        raise Exception("No data found in EDL file")
                    logger.info(f"Parsed {len(edl_data)} shots from EDL")

            if progress_callback:
                progress_callback(0.2, "Analyzing video file...")

            frame_rate = self.video_analyzer.get_video_frame_rate(video_path)
            if frame_rate is None:
                raise Exception("Failed to detect video frame rate")

            logger.info(f"Detected frame rate: {frame_rate} fps")

            if progress_callback:
                progress_callback(0.3, "Calculating durations...")

            tc_handler = TimecodeHandler(frame_rate)
            valid_shots = []
            skipped_shots = []

            for i, shot in enumerate(edl_data, 1):
                try:
                    clip_name = shot.get('Clip Name', '') or ''
                    if clip_name.lower().strip() == 'black video':
                        skipped_shots.append(i)
                        continue

                    if 'Src Start' in shot and 'Src End' in shot:
                        src_start = shot['Src Start']
                        src_end = shot['Src End']

                        if not src_start or not src_end or src_start == 'B' or src_end == 'B':
                            skipped_shots.append(i)
                            continue

                        duration = tc_handler.calculate_duration(src_start, src_end)
                        shot['Duration'] = duration
                        valid_shots.append(shot)
                    else:
                        skipped_shots.append(i)
                except Exception as e:
                    logger.error(f"Error processing shot {i}: {e}")
                    skipped_shots.append(i)
                    continue

            if not valid_shots:
                raise Exception("No valid shots found in EDL after filtering")

            edl_data = valid_shots

            successful_screenshots = 0
            screenshot_dir = os.path.join(output_dir, 'screenshots')
            total_shots = len(edl_data)

            if not disable_screenshots:
                if progress_callback:
                    progress_callback(0.4, "Generating screenshots...")
                os.makedirs(screenshot_dir, exist_ok=True)
                for i, shot in enumerate(edl_data, 1):
                    if progress_callback:
                        progress = 0.4 + (0.4 * i / total_shots)
                        progress_callback(progress, f"Generating screenshot {i}/{total_shots}")

                    # Use exact frame numbers if available (from XML parsing)
                    start_frame = shot.get('_start_frame')
                    end_frame = shot.get('_end_frame')

                    success = self.screenshot_generator.capture_screenshot(
                        video_path,
                        shot.get('Rec Start', shot.get('Src Start', '00:00:00:00')),
                        shot.get('Rec End', shot.get('Src End', '00:00:00:00')),
                        i,
                        shot.get('Clip Name', ''),
                        frame_rate,
                        screenshot_dir,
                        start_frame=start_frame,
                        end_frame=end_frame,
                        width=screenshot_width,
                        height=screenshot_height,
                        quality=screenshot_quality
                    )
                    if success:
                        successful_screenshots += 1
                logger.info(f"Generated {successful_screenshots}/{total_shots} screenshots")

                # Log summary for verification
                if successful_screenshots < total_shots:
                    failed_count = total_shots - successful_screenshots
                    logger.warning(f"Screenshot generation summary: {failed_count} screenshots failed to generate")
            else:
                logger.info("Screenshot generation disabled by user setting")

            if progress_callback:
                progress_callback(0.9, "Updating Excel template...")

            if edl_path:
                base_name = Path(edl_path).stem
            elif show_name:
                base_name = show_name.replace(' ', '_')
            else:
                base_name = Path(video_path).stem
            output_excel = os.path.join(output_dir, f'{base_name}_shotlist.xlsx')

            final_excel = self.excel_updater.update_template(
                template_path, edl_data, mappings, screenshot_dir, output_excel
            )

            if progress_callback:
                progress_callback(1.0, "Processing complete!")

            results['files'] = {
                'excel': final_excel,
            }
            if not disable_screenshots:
                results['files']['screenshots'] = screenshot_dir
            results['files']['csv'] = os.path.join(output_dir, f'{base_name}_shotlist.csv')

            screenshot_note = f" with {successful_screenshots} screenshots" if not disable_screenshots else " (screenshots disabled)"
            results['message'] = f"Successfully processed {len(edl_data)} shots{screenshot_note}"

            return results

        except Exception as e:
            logger.error(f"Processing failed: {e}")
            results['status'] = 'error'
            results['message'] = str(e)
            return results
